import openpyxl
import json
import re

def clean_text(text):
    """Clean text while preserving Arabic characters and formatting"""
    if not text:
        return ""
    # Convert to string and strip whitespace
    text = str(text).strip()
    # Remove extra spaces while preserving Arabic text
    text = ' '.join(filter(None, text.split()))
    # Keep Arabic characters, numbers, and basic punctuation
    text = re.sub(r'[^\u0600-\u06FF\s\d\.,\-?\":;()\w]+', ' ', text)  # Added \w to keep English letters
    
    # Move colon to the left side for Arabic text
    if ':' in text:
        text = text.replace(':', '')  # Remove the original colon
        text = ':' + text  # Add colon at the start
    
    return text

def clean_category(text):
    """Special cleaning for category text which is in English"""
    if not text:
        return ""
    # Just clean up whitespace for English categories
    return str(text).strip()

def is_cell_bold(cell):
    """Check if an Excel cell has bold formatting"""
    if cell.font and cell.font.bold:
        return True
    return False

def extract_ar_questions(xlsx_path):
    questions = []
    processed_ids = set()
    
    # Load workbook with openpyxl to access formatting
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active
    
    row_idx = 1  # Excel rows start at 1
    while row_idx <= sheet.max_row:
        # Check if this is a question row (has an ID in first column)
        # For Arabic, columns are reversed: D=1, C=2, B=3, A=4
        id_cell = sheet.cell(row=row_idx, column=4)  # Column D
        id_text = str(id_cell.value or '')
        id_match = re.match(r'^\s*(\d+)\s*$', id_text)
        
        if id_match:
            question_id = int(id_match.group(1))
            
            # Skip if already processed
            if question_id in processed_ids:
                row_idx += 1
                continue
            
            category = clean_category(str(sheet.cell(row=row_idx, column=3).value))  # Column C - using clean_category
            question_text = "" if 121 <= question_id <= 221 else clean_text(str(sheet.cell(row=row_idx, column=2).value))  # Column B
            
            # Debug print for category
            print(f"Raw category value: '{sheet.cell(row=row_idx, column=3).value}'")
            print(f"Cleaned category: '{category}'")
            
            answers = []
            answer_cells = []
            
            # Get all three answers from column A
            for i in range(3):
                if row_idx + i <= sheet.max_row:
                    answer_cell = sheet.cell(row=row_idx + i, column=1)  # Column A
                    answer_text = clean_text(str(answer_cell.value))
                    if answer_text and answer_text not in answers:
                        answers.append(answer_text)
                        answer_cells.append(answer_cell)
            
            if ((question_text or 121 <= question_id <= 221) and 
                len(answers) == 3 and category):  # Make sure category is not empty
                
                # Find the bold answer
                correct_answer = None
                for idx, (answer, cell) in enumerate(zip(answers, answer_cells)):
                    # Debug print
                    print(f"\nQuestion {question_id}")
                    print(f"Answer {idx + 1}: {answer}")
                    print(f"Bold: {is_cell_bold(cell)}")
                    
                    if is_cell_bold(cell):
                        correct_answer = answer
                        break
                
                # If no bold answer found, use first answer as fallback
                if not correct_answer:
                    correct_answer = answers[0]
                    print(f"Warning: No bold answer found for question {question_id}")
                
                questions.append({
                    "id": question_id,
                    "category": category,
                    "question": question_text,
                    "answers": answers,
                    "correctAnswer": correct_answer
                })
                processed_ids.add(question_id)
            
            row_idx += 3  # Skip the next two rows since we've processed them
        else:
            row_idx += 1
    
    workbook.close()
    questions.sort(key=lambda x: x['id'])
    return questions 