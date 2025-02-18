import pandas as pd
import openpyxl
import json
import re

def clean_text(text):
    if not text:
        return ""
    cleaned = ' '.join(str(text).split()).strip()
    cleaned = re.sub(r'[^\w\s,.-?\'\":;()]', ' ', cleaned)
    return cleaned

def is_cell_bold(cell):
    """Check if an Excel cell has bold formatting"""
    if cell.font and cell.font.bold:
        return True
    return False

def extract_en_questions(xlsx_path):
    questions = []
    processed_ids = set()
    
    # Load workbook with openpyxl to access formatting
    workbook = openpyxl.load_workbook(xlsx_path)
    sheet = workbook.active
    
    row_idx = 1  # Excel rows start at 1
    while row_idx <= sheet.max_row:
        # Check if this is a question row (has an ID in first column)
        id_cell = sheet.cell(row=row_idx, column=1)
        id_text = str(id_cell.value or '')
        id_match = re.match(r'^\s*(\d+)\s*$', id_text)
        
        if id_match:
            question_id = int(id_match.group(1))
            
            # Skip if already processed
            if question_id in processed_ids:
                row_idx += 1
                continue
            
            category = clean_text(str(sheet.cell(row=row_idx, column=2).value))
            question_text = "" if 121 <= question_id <= 221 else clean_text(str(sheet.cell(row=row_idx, column=3).value))
            
            answers = []
            answer_cells = []
            
            # Get all three answers
            for i in range(3):
                if row_idx + i <= sheet.max_row:
                    answer_cell = sheet.cell(row=row_idx + i, column=4)
                    answer_text = clean_text(str(answer_cell.value))
                    if answer_text and answer_text not in answers:
                        answers.append(answer_text)
                        answer_cells.append(answer_cell)
            
            if ((question_text or 121 <= question_id <= 221) and 
                len(answers) == 3 and category):
                
                # Find the bold answer
                correct_answer = None
                for idx, (answer, cell) in enumerate(zip(answers, answer_cells)):
                    # Debug print
                    print(f"\nQuestion {question_id}")
                    print(f"Answer {idx + 1}: {answer}")
                    print(f"Bold: {is_cell_bold(cell)}")
                    
                    if is_cell_bold(cell):
                        correct_answer = answer
                        break
                
                # If no bold answer found, use first answer as fallback
                if not correct_answer:
                    correct_answer = answers[0]
                    print(f"Warning: No bold answer found for question {question_id}")
                
                questions.append({
                    "id": question_id,
                    "category": category,
                    "question": question_text,
                    "answers": answers,
                    "correctAnswer": correct_answer
                })
                processed_ids.add(question_id)
            
            row_idx += 3  # Skip the next two rows since we've processed them
        else:
            row_idx += 1
    
    workbook.close()
    questions.sort(key=lambda x: x['id'])
    return questions 